#!/usr/bin/env python3
"""
fetch_perm_history.py - keep perm_history.json current from DOL's PERM disclosure files.

WHY THIS EXISTS
    perm_history.json drives the two PERM Processing Time charts on tools.html and nothing
    wrote it. Like vb_history.json before it, it was hand-built and then drifted: its last
    entry is FY2026Q2 (Jan-Mar 2026) while FY2026Q3 has been published by DOL, so the chart
    silently understates how far processing times have moved. Nothing in the repo would have
    caught that, which is the actual problem this file fixes.

    Verified 2026-09-03: DOL has published PERM_Disclosure_Data_FY2026_Q3.xlsx covering
    determinations issued 1 Oct 2025 through 30 Jun 2026.

WHY THERE IS A MANUAL DOWNLOAD STEP
    dol.gov returns HTTP 403 to scripted clients. Verified against three URL forms for the
    FY2026 Q3 file, including the /media/ path and the quarter-specific
    /sites/dolgov/files/ETA/oflc/pdfs/FY26Q3/ path - all 403 with a browser User-Agent. A
    real browser downloads them fine. Defeating a federal site's bot protection is out of
    bounds, so this follows the same human-in-the-loop shape as bulletin_pdf_fetch.py:

        1. python3 automation/fetch_perm_history.py --check      # what is missing, and the URL
        2. open that URL in a browser, save the .xlsx
        3. python3 automation/fetch_perm_history.py --parse ~/Downloads/PERM_...xlsx
        4. ...--parse <file> --commit                            # writes perm_history.json

THE FILES ARE CUMULATIVE, WHICH IS THE USEFUL PART
    Each quarterly release covers the whole fiscal year to date, not just that quarter. So one
    download yields every quarter of that fiscal year, and quarters are derived by bucketing
    rows on DECISION_DATE rather than by differencing releases. It also means the file contains
    quarters we ALREADY have, which is what makes the methodology self-check below possible.

THE METHODOLOGY SELF-CHECK, WHICH IS THE POINT OF THE DESIGN
    I did not write the original 14 entries and have no record of how they were computed. The
    obvious failure mode is that a new updater computes "certified" or "median_days" slightly
    differently - say by including Certified-Expired - and appends numbers that are not
    comparable to the existing ones. On a line chart that reads as real-world movement, not as
    a methodology change, and nobody would ever notice.

    So before writing anything, this recomputes the quarters that ALREADY exist in the file and
    compares. It tries both plausible definitions of "certified" and reports which one, if
    either, reproduces the stored values. It refuses to write unless a definition matches, so a
    silent discontinuity cannot be introduced.

WHAT IS NO LONGER AVAILABLE
    The first 8 entries carry median_days_india / _china / _row. Those cannot be reproduced. The
    PERM form changed on 1 June 2023 and the new disclosure layout contains NO foreign-worker
    citizenship or country-of-birth field at all - verified against the official FY2026 Q3
    record layout, whose only country fields are EMP_COUNTRY, EMP_POC_COUNTRY and
    ATTY_AG_COUNTRY, i.e. employer and attorney addresses. Those fields are therefore left off
    new records rather than guessed at, and their absence is a fact about the source rather
    than a gap in this tool.

STATUS OF THIS FILE
    The gap/staleness audit is verified: it runs against the committed JSON.
    The xlsx parser is written against the official record layout but has NOT yet been run on a
    real disclosure file, because dol.gov will not serve one to a script and none was on disk.
    --parse prints what it detected before it computes anything, and --commit is gated on the
    self-check above, so the first real run is diagnosable rather than silently wrong.

Needs openpyxl for --parse (present in automation/.venv). Everything else is stdlib.
"""

import argparse
import datetime
import json
import os
import re
import statistics
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
HISTORY = os.path.join(REPO, "perm_history.json")

# Column names, verbatim from the official record layout (PERM Disclosure Data File
# Structure, FY2026 Q3). Matched by HEADER NAME, never by position: the layout has 137 fields
# and gains and loses some between form revisions, so any positional assumption would silently
# read the wrong column the first time DOL reorders anything.
COL_STATUS = "CASE_STATUS"
COL_RECEIVED = "RECEIVED_DATE"
COL_DECISION = "DECISION_DATE"
COL_CASE = "CASE_NUMBER"
REQUIRED = (COL_STATUS, COL_RECEIVED, COL_DECISION)

# Valid CASE_STATUS values per the layout: "Certified", "Certified-Expired", "Denied",
# "Withdrawn". Whether the site's "certified" series includes Certified-Expired is exactly what
# the self-check resolves, so both readings are computed.
#
# The layout spells it "Certified-Expired" but the FY2026 Q3 data actually emits
# "Certified - Expired", with spaces around the hyphen. Comparing against the un-spaced
# literal silently classified all 16,287 of those rows as "other", which made the
# with-expired reading identical to the strict one and hid the real answer. Statuses are
# therefore normalised before comparison rather than matched verbatim.
CERTIFIED_STRICT = ("certified",)
CERTIFIED_WITH_EXPIRED = ("certified", "certified-expired")


def norm_status(v):
    """CASE_STATUS -> a canonical lowercase form. Collapses whitespace and any spacing
    around hyphens, so "Certified - Expired" and "Certified-Expired" both become
    "certified-expired"."""
    if v is None:
        return ""
    return re.sub(r"\s*-\s*", "-", " ".join(str(v).split()).lower())


CANDIDATE_URLS = [
    "https://www.dol.gov/media/PERM_Disclosure_Data_FY%d_Q%d.xlsx",
    "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/FY%02dQ%d/"
    "PERM_Disclosure_Data_FY%d_Q%d.xlsx",
    "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/"
    "PERM_Disclosure_Data_FY%d_Q%d.xlsx",
]
PERFORMANCE_PAGE = "https://www.dol.gov/agencies/eta/foreign-labor/performance"


# ---------------------------------------------------------------------------
# fiscal quarters
# ---------------------------------------------------------------------------
def fiscal_quarter(d):
    """A date -> ('FY2026Q3'). The federal fiscal year starts 1 October."""
    q = ((d.month - 10) % 12) // 3 + 1
    fy = d.year + 1 if d.month >= 10 else d.year
    return "FY%dQ%d" % (fy, q)


def quarter_key(label):
    """'FY2026Q3' -> (2026, 3), for sorting and comparison."""
    return int(label[2:6]), int(label[7])


def quarter_end(label):
    fy, q = quarter_key(label)
    month = [12, 3, 6, 9][q - 1]
    year = fy - 1 if q == 1 else fy
    day = [31, 31, 30, 30][q - 1]
    return datetime.date(year, month, day)


def quarters_between(a, b):
    out, cur = [], quarter_key(a)
    end = quarter_key(b)
    while cur <= end:
        out.append("FY%dQ%d" % cur)
        cur = (cur[0], cur[1] + 1) if cur[1] < 4 else (cur[0] + 1, 1)
    return out


def latest_expected_quarter(today):
    """The newest quarter DOL should plausibly have published.

    A quarter ends, then DOL takes some weeks to publish. Allowing a full quarter of lag means
    this reports "behind" only when it really is, rather than crying stale every October.
    """
    cur = fiscal_quarter(today)
    fy, q = quarter_key(cur)
    prev = (fy, q - 1) if q > 1 else (fy - 1, 4)
    return "FY%dQ%d" % prev


def urls_for(label):
    fy, q = quarter_key(label)
    return [
        CANDIDATE_URLS[0] % (fy, q),
        CANDIDATE_URLS[1] % (fy % 100, q, fy, q),
        CANDIDATE_URLS[2] % (fy, q),
    ]


# ---------------------------------------------------------------------------
# history file
# ---------------------------------------------------------------------------
def load_history():
    with open(HISTORY, encoding="utf-8") as fh:
        return json.load(fh)


def audit(records):
    problems = []
    labels = [r["quarter"] for r in records]
    if labels != sorted(labels, key=quarter_key):
        problems.append("quarters are not in order")
    dupes = sorted(set(x for x in labels if labels.count(x) > 1))
    if dupes:
        problems.append("duplicate quarter(s): %s" % ", ".join(dupes))
    span = quarters_between(labels[0], labels[-1]) if labels else []
    missing = [q for q in span if q not in set(labels)]
    if missing:
        problems.append("interior gap(s): %s" % ", ".join(missing))
    for r in records:
        for f in ("certified", "denied", "withdrawn", "median_days"):
            if not isinstance(r.get(f), int):
                problems.append("%s: %s is %r, expected an int" % (r["quarter"], f, r.get(f)))
    return problems, {"first": labels[0] if labels else None,
                      "last": labels[-1] if labels else None,
                      "count": len(labels), "expected": len(span)}


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------
def open_rows(path):
    """Stream rows from the disclosure workbook.

    read_only because these files carry six figures of rows; loading one eagerly would be
    gigabytes of cells. data_only so any cached formula value comes through as a value.
    """
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "openpyxl is required for --parse.\n"
            "  automation/.venv/bin/python3 automation/fetch_perm_history.py --parse ...\n"
            "(Homebrew Python is externally managed, PEP 668, so the venv is the way in.)")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        raise SystemExit("%s: the first sheet is empty" % path)
    names = [str(h).strip().upper() if h is not None else "" for h in header]
    idx = {n: i for i, n in enumerate(names) if n}
    missing = [c for c in REQUIRED if c not in idx]
    if missing:
        raise SystemExit(
            "%s: could not find column(s) %s.\n"
            "Columns present (%d): %s\n"
            "The layout may have changed; match by name, do not guess a position."
            % (path, ", ".join(missing), len(names), ", ".join(names[:25]) + " ..."))
    return wb, ws, it, idx, names


def as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(s.split(" ")[0], fmt).date()
        except ValueError:
            continue
    return None


def parse_disclosure(path, verbose=False):
    """-> {quarter: {counts and both median variants}}, plus a scan summary."""
    wb, ws, it, idx, names = open_rows(path)
    print("  detected %d columns; using %s, %s, %s"
          % (len(names), COL_STATUS, COL_RECEIVED, COL_DECISION))

    buckets = {}
    scanned = unparseable_decision = unparseable_received = 0
    statuses = {}

    for row in it:
        scanned += 1
        raw_status = row[idx[COL_STATUS]]
        status = norm_status(raw_status)
        # Report the value VERBATIM as DOL wrote it, and bucket on the normalised form. The
        # raw spelling is what makes a status mismatch diagnosable at all, so it is not
        # normalised away in the printout.
        seen = (str(raw_status).strip() if raw_status is not None else "")
        statuses[seen] = statuses.get(seen, 0) + 1
        dec = as_date(row[idx[COL_DECISION]])
        if dec is None:
            unparseable_decision += 1
            continue
        q = fiscal_quarter(dec)
        b = buckets.setdefault(q, {"certified": 0, "certified_expired": 0, "denied": 0,
                                   "withdrawn": 0, "other": 0,
                                   "days_strict": [], "days_with_expired": []})
        if status == "certified":
            b["certified"] += 1
        elif status == "certified-expired":
            b["certified_expired"] += 1
        elif status == "denied":
            b["denied"] += 1
        elif status == "withdrawn":
            b["withdrawn"] += 1
        else:
            b["other"] += 1

        if status in CERTIFIED_WITH_EXPIRED:
            rec = as_date(row[idx[COL_RECEIVED]])
            if rec is None:
                unparseable_received += 1
            else:
                days = (dec - rec).days
                if days >= 0:                      # a negative span is a data error, not a wait
                    b["days_with_expired"].append(days)
                    if status == "certified":
                        b["days_strict"].append(days)

    wb.close()
    summary = {"rows": scanned, "no_decision_date": unparseable_decision,
               "no_received_date": unparseable_received, "statuses": statuses}
    return buckets, summary


def records_from(buckets, include_expired):
    """Turn buckets into perm_history-shaped records under one 'certified' definition."""
    out = []
    for q in sorted(buckets, key=quarter_key):
        b = buckets[q]
        days = b["days_with_expired"] if include_expired else b["days_strict"]
        cert = b["certified"] + (b["certified_expired"] if include_expired else 0)
        if not days:
            continue
        out.append({
            "quarter": q,
            "certified": cert,
            "denied": b["denied"],
            "withdrawn": b["withdrawn"],
            "median_days": int(round(statistics.median(days))),
            "certified_with_dates": len(days),
        })
    return out


# A later release may legitimately RESTATE an earlier quarter: a denial that is reopened and
# redecided gets a new DECISION_DATE and leaves the quarter it was first counted in. Observed
# on FY2026Q2, where the newer file has 146 fewer denials and 1 fewer certification than the
# stored series - 0.35% of that quarter's 41,987 decisions - while `withdrawn` (a terminal
# status) and the median are untouched.
#
# Demanding exact equality would therefore make the gate unpassable forever on any revised
# quarter, which defeats its purpose: it exists to catch a METHODOLOGY change, not to freeze
# DOL's data. So the test is split in two:
#
#   * median_days must match EXACTLY. It is the methodology fingerprint. Getting the
#     "certified" definition wrong moves it enormously - certified-only gives 783 days for
#     FY2026Q1 against a stored 498 - because it changes which population is measured, not
#     just how many rows are in it.
#   * the counts may drift by at most TOLERANCE_SHARE of that quarter's total decisions.
#     Expressed against the quarter's caseload rather than per-field on purpose: a revision
#     moves a small number of CASES, and 146 cases is 0.35% of the quarter but 10.5% of the
#     `denied` field alone, so a per-field percentage would reject an ordinary restatement.
#
# The separation is wide. FY2026Q1 under the wrong definition is off by 16,163 certifications
# against a tolerance of about 90, so a methodology error cannot slip through as drift.
TOLERANCE_SHARE = 0.005


def compare(existing, computed):
    """Diff the quarters both sides have.

    Returns (verdict, rows) where verdict is "exact", "within" (medians exact, counts inside
    the revision tolerance) or "mismatch". rows are (quarter, field, stored, computed,
    allowed) for every field that differs, so an exact match can still be reported.
    """
    have = {r["quarter"]: r for r in existing}
    rows = []
    any_overlap = False
    exact = True
    within = True
    for c in computed:
        e = have.get(c["quarter"])
        if not e:
            continue
        any_overlap = True
        total = sum(int(e.get(f) or 0) for f in ("certified", "denied", "withdrawn"))
        allowed = max(1, int(round(total * TOLERANCE_SHARE)))
        for f in ("certified", "denied", "withdrawn", "median_days"):
            was, now = e.get(f), c.get(f)
            if was == now:
                continue
            exact = False
            # The median carries no tolerance; a count is judged against the quarter's size.
            budget = 0 if f == "median_days" else allowed
            if was is None or now is None or abs(int(was) - int(now)) > budget:
                within = False
            rows.append((c["quarter"], f, was, now, budget))
    if not any_overlap:
        return "mismatch", rows
    if exact:
        return "exact", rows
    return ("within" if within else "mismatch"), rows


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def cmd_check(records, today):
    problems, info = audit(records)
    print("perm_history.json audit")
    print("  quarters        : %d (%s -> %s)" % (info["count"], info["first"], info["last"]))
    print("  expected span   : %d" % info["expected"])
    want = latest_expected_quarter(today)
    print("  today           : %s (fiscal %s)" % (today.isoformat(), fiscal_quarter(today)))
    print("  newest expected : %s" % want)

    behind = []
    if info["last"] and quarter_key(info["last"]) < quarter_key(want):
        behind = quarters_between(info["last"], want)[1:]

    if behind:
        print("\n  STALE by %d quarter(s): %s" % (len(behind), ", ".join(behind)))
        print("\n  Download the newest cumulative file - it covers the whole fiscal year, so one")
        print("  file supplies every missing quarter in that year:")
        for u in urls_for(behind[-1]):
            print("      %s" % u)
        print("\n  dol.gov 403s scripted clients, so open it in a browser, then:")
        print("      automation/.venv/bin/python3 automation/fetch_perm_history.py \\")
        print("          --parse ~/Downloads/PERM_Disclosure_Data_%s.xlsx"
              % behind[-1].replace("Q", "_Q"))
        print("  Index of releases: %s" % PERFORMANCE_PAGE)
    else:
        print("\n  Up to date.")

    if problems:
        print("\n  problems:")
        for p in problems:
            print("    - %s" % p)
    return 1 if (problems or behind) else 0


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--check", action="store_true", help="audit and report staleness (default)")
    ap.add_argument("--verify", action="store_true", help="structural audit only")
    ap.add_argument("--urls", metavar="FYnnnnQn", help="print candidate URLs for a quarter")
    ap.add_argument("--parse", metavar="XLSX", help="parse a downloaded disclosure file")
    ap.add_argument("--commit", action="store_true", help="write perm_history.json")
    ap.add_argument("--force", action="store_true",
                    help="write even if the methodology self-check fails (records why)")
    ap.add_argument("--date", metavar="YYYY-MM-DD", help="override today, for testing")
    args = ap.parse_args()

    today = (datetime.date.fromisoformat(args.date) if args.date else datetime.date.today())
    records = load_history()

    if args.urls:
        for u in urls_for(args.urls):
            print(u)
        return 0

    if args.verify:
        problems, info = audit(records)
        if problems:
            print("FAILED: %d problem(s)" % len(problems))
            for p in problems:
                print("  - %s" % p)
            return 1
        print("verified: %d quarters, %s -> %s, no gaps or duplicates"
              % (info["count"], info["first"], info["last"]))
        return 0

    if not args.parse:
        return cmd_check(records, today)

    # ---- parse ----
    print("parsing %s" % args.parse)
    buckets, summary = parse_disclosure(args.parse)
    print("  scanned %d rows; %d had no usable DECISION_DATE" %
          (summary["rows"], summary["no_decision_date"]))
    print("  CASE_STATUS values seen: %s" %
          ", ".join("%s=%d" % kv for kv in sorted(summary["statuses"].items(),
                                                  key=lambda kv: -kv[1])))
    if summary["no_received_date"]:
        print("  %d certified rows had no usable RECEIVED_DATE and are excluded from the median"
              % summary["no_received_date"])

    strict = records_from(buckets, include_expired=False)
    loose = records_from(buckets, include_expired=True)
    print("\n  quarters found: %s" % ", ".join(r["quarter"] for r in strict))

    # ---- methodology self-check ----
    v_strict, diff_strict = compare(records, strict)
    v_loose, diff_loose = compare(records, loose)
    overlap = [r["quarter"] for r in strict if r["quarter"] in {x["quarter"] for x in records}]
    print("\n  methodology self-check against %d quarter(s) already in the file: %s"
          % (len(overlap), ", ".join(overlap) or "none"))

    STRICT_LABEL = '"certified" excludes Certified-Expired'
    LOOSE_LABEL = '"certified" includes Certified-Expired'

    def show(name, rows):
        print("    --- %s, differences ---" % name)
        for q, f, was, now, budget in rows[:8]:
            note = "" if budget == 0 else "  (tolerance %+d)" % budget
            print("        %s %-20s stored=%s computed=%s%s" % (q, f, was, now, note))

    chosen = None
    if not overlap:
        print("    NO OVERLAP - cannot confirm the new numbers are comparable to the old ones.")
    # Prefer an exact match on either definition, then a within-tolerance one. Checking both
    # exacts before either tolerance keeps a definition that reproduces the series perfectly
    # from losing to one that merely lands close.
    elif v_strict == "exact":
        chosen = strict
        print("    MATCH, exact: %s" % STRICT_LABEL)
    elif v_loose == "exact":
        chosen = loose
        print("    MATCH, exact: %s" % LOOSE_LABEL)
    elif v_strict == "within" or v_loose == "within":
        if v_strict == "within":
            chosen, label, rows = strict, STRICT_LABEL, diff_strict
        else:
            chosen, label, rows = loose, LOOSE_LABEL, diff_loose
        print("    MATCH on methodology: %s" % label)
        print("    Every median is exact and every count is inside the revision tolerance,")
        print("    so this is DOL restating an earlier quarter, not a different definition.")
        show("accepted", rows)
    else:
        print("    NEITHER definition reproduces the stored values.")
        for name, rows in (("strict", diff_strict), ("with-expired", diff_loose)):
            show(name, rows)

    src = chosen if chosen else strict
    print("\n  computed records:")
    for r in src:
        mark = "  (already in file)" if r["quarter"] in {x["quarter"] for x in records} else "  NEW"
        print("    %-10s certified=%-7d denied=%-6d withdrawn=%-6d median_days=%-4d%s"
              % (r["quarter"], r["certified"], r["denied"], r["withdrawn"],
                 r["median_days"], mark))

    have = {r["quarter"] for r in records}
    new = [r for r in src if r["quarter"] not in have]
    if not new:
        print("\n  Nothing new to add.")
        return 0
    print("\n  %d new quarter(s): %s" % (len(new), ", ".join(r["quarter"] for r in new)))
    print("  Note: median_days_india / _china / _row are NOT produced. The post-June-2023")
    print("  PERM form carries no foreign-worker citizenship field, so they cannot be.")

    if not args.commit:
        print("\nDry run: nothing written. Re-run with --commit.")
        return 0
    if chosen is None and not args.force:
        print("\nREFUSING TO WRITE: the self-check did not confirm a matching methodology, so")
        print("these numbers may not be comparable to the existing series. A discontinuity on")
        print("the chart would read as real movement. Re-run with --force to override.")
        return 1

    merged = sorted(records + new, key=lambda r: quarter_key(r["quarter"]))
    problems, info = audit(merged)
    if problems:
        print("\nREFUSING TO WRITE: merged result fails its own audit:")
        for p in problems[:8]:
            print("  - %s" % p)
        return 1
    with open(HISTORY, "w", encoding="utf-8") as fh:
        json.dump(merged, fh, indent=2)
        fh.write("\n")
    print("\nwrote %s" % HISTORY)
    print("verified: %d quarters, %s -> %s, no gaps"
          % (info["count"], info["first"], info["last"]))
    return 0


if __name__ == "__main__":
    sys.exit(main())
